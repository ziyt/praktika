
import os
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

APP_DIR = Path(__file__).resolve().parent.parent
DB_PATH = APP_DIR / "db" / "shoe_store.db"
SCHEMA_PATH = APP_DIR / "db" / "schema.sql"
IMPORT_DIR = Path("/tmp/prakt_rar/import")

ROLES = ["Гость", "Клиент", "Менеджер", "Администратор"]

ROLE_ALIASES = {
    "Авторизированный клиент": "Клиент",
    "Авторизованный клиент":   "Клиент",
}

def create_database(conn: sqlite3.Connection) -> None:
    with open(SCHEMA_PATH, "r", encoding="utf-8") as f:
        conn.executescript(f.read())
    conn.commit()

def upsert_lookup(conn: sqlite3.Connection, table: str, name: str) -> int:
    cur = conn.execute(f"SELECT id FROM {table} WHERE name = ?", (name,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur = conn.execute(f"INSERT INTO {table}(name) VALUES (?)", (name,))
    return cur.lastrowid

def seed_roles(conn: sqlite3.Connection) -> dict:
    role_ids = {}
    for role in ROLES:
        role_ids[role] = upsert_lookup(conn, "roles", role)
    conn.commit()
    return role_ids

def seed_employees(conn: sqlite3.Connection, role_ids: dict) -> dict:
    wb = openpyxl.load_workbook(IMPORT_DIR / "user_import.xlsx", data_only=True)
    ws = wb.active
    inserted = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        role, full_name, login, password = row[:4]
        if not (role and full_name and login and password):
            continue
        role_name = role.strip()
        role_name = ROLE_ALIASES.get(role_name, role_name)
        role_id = role_ids[role_name]
        conn.execute(
            "INSERT INTO users(full_name, login, password, role_id) VALUES (?, ?, ?, ?)",
            (full_name.strip(), login.strip(), password.strip(), role_id),
        )
        inserted[full_name.strip()] = login.strip()
    conn.commit()
    return inserted

def seed_pickup_points(conn: sqlite3.Connection) -> dict:
    wb = openpyxl.load_workbook(IMPORT_DIR / "Пункты выдачи_import.xlsx", data_only=True)
    ws = wb.active
    idx_to_id = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or not row[0]:
            continue
        address = row[0].strip()
        cur = conn.execute("INSERT INTO pickup_points(address) VALUES (?)", (address,))
        idx_to_id[i] = cur.lastrowid
    conn.commit()
    return idx_to_id

def seed_products(conn: sqlite3.Connection) -> None:
    wb = openpyxl.load_workbook(IMPORT_DIR / "Tovar.xlsx", data_only=True)
    ws = wb.active
    for row in list(ws.iter_rows(values_only=True))[1:]:
        if not row or not row[0]:
            continue
        (article, name, unit, price, supplier, manufacturer,
         category, discount, stock, description, photo) = row[:11]
        cat_id = upsert_lookup(conn, "categories", category.strip())
        man_id = upsert_lookup(conn, "manufacturers", manufacturer.strip())
        sup_id = upsert_lookup(conn, "suppliers", supplier.strip())
        unit_id = upsert_lookup(conn, "units", unit.strip())
        photo_path = f"resources/products/{photo.strip()}" if photo else None
        conn.execute(
            """INSERT INTO products(article, name, description, photo_path,
                                    price, discount, stock_qty,
                                    category_id, manufacturer_id, supplier_id, unit_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (article.strip(), name.strip(), (description or "").strip(), photo_path,
             float(price), int(discount or 0), int(stock or 0),
             cat_id, man_id, sup_id, unit_id),
        )
    conn.commit()

def parse_date(value):
    if value is None or value == "":
        return None
    if hasattr(value, "date"):
        return value.date().isoformat()
    return str(value).strip()

def parse_order_items(s: str) -> list:
    parts = [p.strip() for p in s.split(",") if p.strip()]
    items = []
    for i in range(0, len(parts) - 1, 2):
        article = parts[i]
        try:
            qty = int(parts[i + 1])
        except ValueError:
            continue
        items.append((article, qty))
    return items

def ensure_client(conn: sqlite3.Connection, full_name: str,
                  existing: dict, client_role_id: int, counter: list) -> int:
    """Создаёт клиента по ФИО (если не нашёлся среди уже импортированных)."""
    cur = conn.execute("SELECT id FROM users WHERE full_name = ?", (full_name,))
    row = cur.fetchone()
    if row:
        return row[0]
    counter[0] += 1
    login = f"client{counter[0]:03d}@shop.local"
    password = f"Cl{counter[0]:04d}!"
    cur = conn.execute(
        "INSERT INTO users(full_name, login, password, role_id) VALUES (?, ?, ?, ?)",
        (full_name, login, password, client_role_id),
    )
    return cur.lastrowid

def seed_orders(conn: sqlite3.Connection, pickup_map: dict, role_ids: dict) -> None:
    wb = openpyxl.load_workbook(IMPORT_DIR / "Заказ_import.xlsx", data_only=True)
    ws = wb.active
    client_counter = [0]
    cache = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        if not row or row[0] is None:
            continue
        order_id = int(row[0])
        items_raw = (row[1] or "").strip()
        order_date = parse_date(row[2])
        delivery_date = parse_date(row[3])
        pickup_idx = int(row[4]) if row[4] is not None else None
        client_name = (row[5] or "").strip()
        pickup_code = int(row[6]) if row[6] is not None else 0
        status = (row[7] or "Новый").strip()

        if pickup_idx not in pickup_map:
            print(f"  ! заказ {order_id}: нет пункта выдачи {pickup_idx}, пропуск")
            continue

        client_id = ensure_client(conn, client_name, cache,
                                  role_ids["Клиент"], client_counter)

        conn.execute(
            """INSERT INTO orders(id, order_date, delivery_date,
                                  pickup_point_id, client_id, pickup_code, status)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (order_id, order_date, delivery_date,
             pickup_map[pickup_idx], client_id, pickup_code, status),
        )
        for article, qty in parse_order_items(items_raw):
            try:
                conn.execute(
                    "INSERT INTO order_items(order_id, product_article, quantity) VALUES (?, ?, ?)",
                    (order_id, article, qty),
                )
            except sqlite3.IntegrityError as e:
                print(f"  ! позиция заказа {order_id} ({article} x{qty}): {e}")
    conn.commit()

def export_sql_dump(conn: sqlite3.Connection) -> None:
    dump_path = APP_DIR / "db" / "shoe_store_dump.sql"
    with open(dump_path, "w", encoding="utf-8") as f:
        for line in conn.iterdump():
            f.write(line + "\n")
    print(f"  → SQL-дамп: {dump_path}")

def main() -> None:
    if DB_PATH.exists():
        DB_PATH.unlink()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")

    print("Создание схемы...")
    create_database(conn)

    print("Импорт ролей...")
    role_ids = seed_roles(conn)

    print("Импорт сотрудников...")
    seed_employees(conn, role_ids)

    print("Импорт пунктов выдачи...")
    pickup_map = seed_pickup_points(conn)

    print("Импорт товаров...")
    seed_products(conn)

    print("Импорт заказов (с автосозданием клиентов)...")
    seed_orders(conn, pickup_map, role_ids)

    print("Экспорт SQL-дампа...")
    export_sql_dump(conn)

    stats = {}
    for t in ["roles", "users", "categories", "manufacturers", "suppliers",
              "units", "products", "pickup_points", "orders", "order_items"]:
        stats[t] = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
    conn.close()

    print("\nГотово. Записей по таблицам:")
    for t, n in stats.items():
        print(f"  {t:<16} {n}")

if __name__ == "__main__":
    main()
